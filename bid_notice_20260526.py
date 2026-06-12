"""
나라장터 자동 조회 스크립트
  - 오전 08:00 실행 → 전일 0000 ~ 전일 2359
  - 발주계획 발주시기: 당월 -2달 ~ +10달
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

import math
import os
import time
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

import pandas as pd
import requests
import schedule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

import logging
logging.getLogger("urllib3").setLevel(logging.ERROR)


# ══════════════════════════════════════════════════════════════════════
# ★ 사용자 설정
# ══════════════════════════════════════════════════════════════════════

SERVICE_KEY = "74dbb62cd9ef2a31b5a4b12af4bba9b6fe3a4e8b8e64a363d36cb0017c6b5d17"

# 엑셀 파일 저장 폴더
OUTPUT_DIR = r"\\10.2.31.100\pub\3. 입찰공고관리\01.입찰공고_자동추출"

ROWS = 50

# demand_org 수요기관
TARGET_ORG = [
    "한국연구재단",
    "정보통신기획평가원",
    "한국과학기술정보연구원",
    "한국과학기술기획평가원",
    "중소기업기술정보진흥원",
    "한국보건산업진흥원",
    "정보통신산업진흥원",
    "한국과학기술연구원",
    "한국산업기술기획평가원",
    "한국지능정보사회진흥원",
]

# service_type 업무구분
TARGET_SERVICE_TYPES = ["일반용역", "기술용역"]

# ══════════════════════════════════════════════════════════════════════
# ★ 수동 날짜 설정
# ══════════════════════════════════════════════════════════════════════

#MANUAL_START = "202605200000"
#MANUAL_END   = "202605202359"


# ══════════════════════════════════════════════════════════════════════
# API 엔드포인트
# ══════════════════════════════════════════════════════════════════════

URL_BID  = "https://apis.data.go.kr/1230000/ad/BidPublicInfoService/getBidPblancListInfoServc"
URL_SPEC = "https://apis.data.go.kr/1230000/ao/HrcspSsstndrdInfoService/getPublicPrcureThngInfoServcPPSSrch"
URL_PLAN = "https://apis.data.go.kr/1230000/ao/OrderPlanSttusService/getOrderPlanSttusListServcPPSSrch"


# ══════════════════════════════════════════════════════════════════════
# 공통 유틸
# ══════════════════════════════════════════════════════════════════════
FONT_NAME    = "맑은 고딕"
HEADER_FILL  = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ROW_FILL_ODD = PatternFill(start_color="EEF2F9", end_color="EEF2F9", fill_type="solid")
THIN_SIDE    = Side(border_style="thin", color="C9C9C9")
CELL_BORDER  = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


def now_kst() -> datetime:
    return datetime.now(ZoneInfo("Asia/Seoul"))


def shift_month(dt: datetime, delta: int) -> str:
    """당월 기준으로 delta개월 이동한 YYYYMM 반환 (delta 음수=과거, 양수=미래)"""
    total = dt.year * 12 + (dt.month - 1) + delta
    y, m = divmod(total, 12)
    return f"{y}{m + 1:02d}"


def fmt_date(yyyymmddhhmi: str) -> str:
    d = yyyymmddhhmi
    return f"{d[:4]}/{d[4:6]}/{d[6:8]}" if len(d) >= 8 else d


def safe_int(value: str, default: int = 0) -> int:
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return default

os.system("chcp 65001")

#로그 설정
LOG_ROOT = r"C:\bid_notice\logs"
log_month = now_kst().strftime("%Y%m")
LOG_DIR = os.path.join(LOG_ROOT, log_month)     #C:\bid_notice\logs\202605
os.makedirs(LOG_DIR, exist_ok=True)

log_time = now_kst().strftime("%Y%m%d_%H%M%S")  #20260512_164233.log
log_file = os.path.join(LOG_DIR, f"{log_time}.log")

logging.basicConfig(
level=logging.INFO,
format="%(asctime)s [%(levelname)s] %(message)s",
handlers=[
logging.FileHandler(log_file, encoding="utf-8"),
logging.StreamHandler()
]
)

logger = logging.getLogger(__name__)


def out_path(run_time: str, filename: str) -> str:
    month_dir = run_time[:6]
    
    save_dir = os.path.join(OUTPUT_DIR, month_dir, run_time)
    os.makedirs(save_dir, exist_ok=True)
    return os.path.join(save_dir, filename)

def make_session() -> requests.Session:
    session = requests.Session()
    retry = Retry(total=3, backoff_factor=1, status_forcelist=[500, 502, 503, 504])
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    return session


def fetch_all_pages(session: requests.Session, url: str, base_params: dict) -> list[ET.Element]:
    """페이지네이션을 순회하며 모든 <item> 요소 반환"""
    params = {**base_params, "pageNo": 1}
    resp = session.get(url, params=params, timeout=(3, 15))
    resp.raise_for_status()

    root = ET.fromstring(resp.content)
    total_count = int(root.findtext(".//totalCount", default="0"))
    total_pages = math.ceil(total_count / base_params.get("numOfRows", 50))

    logger.info(f"  전체 건수: {total_count}  /  총 페이지: {total_pages}")

    all_items: list[ET.Element] = []

    for page in range(1, total_pages + 1):
        logger.info(f"  [{page}/{total_pages}] 페이지 조회중...")
        params = {**base_params, "pageNo": page}

        try:
            resp = session.get(url, params=params, timeout=(3, 15))
            resp.raise_for_status()
            root  = ET.fromstring(resp.content)
            items = root.findall(".//item")

            if not items:
                print("  더 이상 데이터 없음, 순회 종료")
                break

            all_items.extend(items)

        except Exception as e:
            logger.info(f"  {page} 페이지 에러: {e}")

        time.sleep(0.1)

    return all_items


def write_meta(ws, cell: str, label: str, value) -> None:
    next_col = chr(ord(cell[0]) + 1)
    row = cell[1:]
    ws[cell].value = label
    ws[cell].font  = Font(name=FONT_NAME, bold=True, size=9, color="555555")
    ws[f"{next_col}{row}"].value = value
    ws[f"{next_col}{row}"].font  = Font(name=FONT_NAME, size=9)


def apply_excel_style(ws, start_row: int, column_widths: dict, number_cols: list = None) -> None:
    header_row = start_row + 1

    for cell in ws[header_row]:
        cell.font      = Font(name=FONT_NAME, bold=True, size=9, color="FFFFFF")
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = CELL_BORDER
    ws.row_dimensions[header_row].height = 30

    for i, row in enumerate(ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row)):
        fill = ROW_FILL_ODD if i % 2 == 0 else PatternFill()
        for cell in row:
            cell.font      = Font(name=FONT_NAME, size=9)
            cell.fill      = fill
            cell.border    = CELL_BORDER
            cell.alignment = Alignment(vertical="center")

    if number_cols:
        for col_letter in number_cols:
            col_idx = ord(col_letter.upper()) - 64
            for row in ws.iter_rows(
                min_row=header_row + 1, max_row=ws.max_row,
                min_col=col_idx, max_col=col_idx,
            ):
                for cell in row:
                    cell.number_format = "#,##0"
                    cell.alignment     = Alignment(horizontal="right", vertical="center")

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


# ══════════════════════════════════════════════════════════════════════
# 01. 입찰공고
# ══════════════════════════════════════════════════════════════════════

def fetch_입찰공고(session, start_date: str, end_date: str) -> pd.DataFrame:
    logger.info("\n===== 01. 입찰공고 조회 시작 =====")

    base_params = {
        "serviceKey": SERVICE_KEY,
        "numOfRows":  ROWS,
        "inqryDiv":   1,
        "inqryBgnDt": start_date,
        "inqryEndDt": end_date,
    }
    items = fetch_all_pages(session, URL_BID, base_params)
    
    latest_items = {}

    for item in items:
        bid_no = item.findtext("bidNtceNo", "")
        detail_bid_no = item.findtext("bidNtceOrd", "0")

        try:
            ord_no = int(detail_bid_no)
        except ValueError:
            ord_no = 0

        if bid_no not in latest_items:
            latest_items[bid_no] = item
        else:
            prev_ord = int(
                latest_items[bid_no].findtext("bidNtceOrd", "0")
            )

            if ord_no > prev_ord:
                latest_items[bid_no] = item

    results = []

    for item in latest_items.values():
        
        bid_no = item.findtext("bidNtceNo", "")
        detail_bid_no = item.findtext("bidNtceOrd", "0")
        입찰공고번호 = f"{bid_no}-{detail_bid_no}"
        
        service_type = item.findtext("srvceDivNm", "")
        demand_org   = item.findtext("dminsttNm", "")

        if demand_org not in TARGET_ORG or service_type not in TARGET_SERVICE_TYPES:
            continue

        results.append({
            "NO":           len(results) + 1,
            "업무구분":     service_type,
            "업무여부":     "",
            "구분":         item.findtext("ntceKindNm", ""),
            "입찰공고번호": 입찰공고번호,
            "공고명":       item.findtext("bidNtceNm", ""),
            "공고기관":     item.findtext("ntceInsttNm", ""),
            "수요기관":     demand_org,
            "게시일시":     item.findtext("bidNtceDt", ""),
            "입찰마감일시": item.findtext("bidClseDt", ""),
            #"단계":         "",
            #"세부절차":     "",
            #"세부절차상태": "",
            "계약체결방법명": item.findtext("cntrctCnclsMthdNm", ""),
            "낙찰방법명": item.findtext("sucsfbidMthdNm", ""),
            "배정예산금액": safe_int(item.findtext("asignBdgtAmt", "")),
            "추정가격": safe_int(item.findtext("presmptPrce", "")),
        })

    logger.info(f"  입찰공고 필터 결과: {len(results)}건")
    df = pd.DataFrame(results)

    if not df.empty:
        df = df.sort_values("게시일시", ascending=False).reset_index(drop=True)
        df["NO"] = df.index + 1
    return df


def save_입찰공고(df: pd.DataFrame, start_date: str, end_date: str, run_time: str,) -> str:
    ts        = now_kst().strftime("%Y%m%d_%H%M")
    file_path = out_path(run_time, f"01.입찰공고목록_{ts}.xlsx")
    start_row = 5

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="입찰공고", index=False, startrow=start_row)
        ws = writer.sheets["입찰공고"]

        ws["A1"] = "입찰공고목록"
        ws["A1"].font = Font(name=FONT_NAME, bold=True, size=15)

        write_meta(ws, "A2", "생성일",   now_kst().strftime("%Y/%m/%d %H:%M"))
        write_meta(ws, "A3", "조회기간", f"{fmt_date(start_date)} ~ {fmt_date(end_date)}")
        write_meta(ws, "A4", "총 건수",  f"{len(df)}건")

        apply_excel_style(ws, start_row, {
            "A": 6,  "B": 12, "C": 10, "D": 12, "E": 22,
            "F": 46, "G": 28, "H": 26, "I": 20, "J": 20,
            "K": 12, "L": 28, "M": 16, "N": 16, 
        }, number_cols=["M", "N"])

    logger.info(f"  저장 완료: {file_path}")
    return file_path


# ══════════════════════════════════════════════════════════════════════
# 02. 사전규격공개
# ══════════════════════════════════════════════════════════════════════

def fetch_사전규격(session, start_date: str, end_date: str) -> pd.DataFrame:
    logger.info("\n===== 02. 사전규격공개 조회 시작 =====")

    base_params = {
        "serviceKey": SERVICE_KEY,
        "numOfRows":  ROWS,
        "inqryDiv":   1,
        "inqryBgnDt": start_date,
        "inqryEndDt": end_date,
    }
    items = fetch_all_pages(session, URL_SPEC, base_params)

    results, visited = [], set()

    for item in items:
        spec_no = item.findtext("bfSpecRgstNo", "")
        if spec_no in visited:
            continue
        visited.add(spec_no)

        service_type = item.findtext("bsnsDivNm", "")
        demand_org   = item.findtext("rlDminsttNm", "")

        if demand_org not in TARGET_ORG or service_type not in TARGET_SERVICE_TYPES:
            continue

        results.append({
            "NO":         len(results) + 1,
            "업무구분":   service_type,
            "업무여부":   "",
            "사업명":     item.findtext("prdctClsfcNoNm", ""),
            "수요기관":   demand_org,
            "공고기관":   item.findtext("orderInsttNm", ""),
            "담당자":     item.findtext("ofclNm", ""),
            "진행일자":   item.findtext("rcptDt", ""),
            "진행상태":   "",
            "참조여부":   "",
            "업체등록수": "",
            "예산금액":   safe_int(item.findtext("asignBdgtAmt", "0")),
        })

    logger.info(f"  사전규격 필터 결과: {len(results)}건")
    df = pd.DataFrame(results)
    if not df.empty:
        df = df.sort_values("진행일자", ascending=False).reset_index(drop=True)
        df["NO"] = df.index + 1
    return df


def save_사전규격(df: pd.DataFrame, start_date: str, end_date: str, run_time: str,) -> str:
    ts        = now_kst().strftime("%Y%m%d_%H%M")
    file_path = out_path(run_time, f"02.발주목록_사전규격공개_{ts}.xlsx")
    start_row = 8

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="사전규격공개", index=False, startrow=start_row)
        ws = writer.sheets["사전규격공개"]

        ws["A1"] = "발주목록_사전규격공개"
        ws["A1"].font = Font(name=FONT_NAME, bold=True, size=15)

        write_meta(ws, "A2", "생성일",   now_kst().strftime("%Y-%m-%d %H:%M"))
        write_meta(ws, "A3", "검색유형", "사전규격공개")
        write_meta(ws, "A4", "수요기관", ", ".join(TARGET_ORG))
        write_meta(ws, "A5", "업무구분", ", ".join(TARGET_SERVICE_TYPES))
        write_meta(ws, "A6", "진행일자", f"{fmt_date(start_date)} ~ {fmt_date(end_date)}")
        write_meta(ws, "A7", "총 건수",  f"{len(df)}건")

        apply_excel_style(ws, start_row, {
            "A": 6,  "B": 13, "C": 10, "D": 45,
            "E": 24, "F": 24, "G": 11, "H": 20,
            "I": 13, "J": 10, "K": 10, "L": 16,
        }, number_cols=["L"])

    logger.info(f"  저장 완료: {file_path}")
    return file_path


# ══════════════════════════════════════════════════════════════════════
# 03. 발주계획
# ══════════════════════════════════════════════════════════════════════

def fetch_발주계획(session, start_date: str, end_date: str,
                  order_start_ym: str, order_end_ym: str) -> pd.DataFrame:
    logger.info("\n===== 03. 발주계획 조회 시작 =====")

    base_params = {
        "serviceKey": SERVICE_KEY,
        "numOfRows":  ROWS,
        "inqryDiv":   1,
        "inqryBgnDt": start_date,
        "inqryEndDt": end_date,
        "orderBgnYm": order_start_ym,
        "orderEndYm": order_end_ym,
    }
    items = fetch_all_pages(session, URL_PLAN, base_params)

    results, visited = [], set()

    for item in items:
        plan_no = item.findtext("orderPlanUntyNo", "")
        if plan_no in visited:
            continue
        visited.add(plan_no)

        service_type = item.findtext("bsnsDivNm", "")
        demand_org   = item.findtext("orderInsttNm", "")
        발주년도      = item.findtext("orderYear", "")
        발주시기      = item.findtext("orderMnth", "")

        if demand_org not in TARGET_ORG:
            continue

        results.append({
            "No":       len(results) + 1,
            "업무구분": service_type,
            "업무여부": "",
            "사업명":   item.findtext("bizNm", ""),
            "수요기관": demand_org,
            "담당자":   item.findtext("ofclNm", ""),
            "진행일자": item.findtext("nticeDt", ""),
            "진행상태": "",
            "참조여부": "",
            "발주시기": f"{발주년도}/{발주시기.zfill(2)}",
            "예산금액": safe_int(item.findtext("sumOrderAmt", "0")),
        })

    logger.info(f"  발주계획 필터 결과: {len(results)}건")
    df = pd.DataFrame(results)
    if not df.empty:
        df = df.sort_values("진행일자", ascending=False).reset_index(drop=True)
        df["No"] = df.index + 1
    return df


def save_발주계획(df: pd.DataFrame, start_date: str, end_date: str, run_time: str,) -> str:
    ts        = now_kst().strftime("%Y%m%d_%H%M")
    file_path = out_path(run_time, f"03.발주목록_발주계획_{ts}.xlsx")
    start_row = 9

    month_list   = sorted(set(df["발주시기"].astype(str))) if not df.empty else []
    service_text = ", ".join(sorted(set(df["업무구분"].astype(str)))) if not df.empty else ""

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="발주계획", index=False, startrow=start_row)
        ws = writer.sheets["발주계획"]

        ws["A1"] = "발주목록_발주계획"
        ws["A1"].font = Font(name=FONT_NAME, bold=True, size=15)

        write_meta(ws, "A2", "생성일",   now_kst().strftime("%Y-%m-%d %H:%M"))
        write_meta(ws, "A3", "검색유형", "발주계획")
        write_meta(ws, "A4", "진행일자", f"{fmt_date(start_date)} ~ {fmt_date(end_date)}")
        write_meta(ws, "A5", "발주시기", ", ".join(month_list))
        write_meta(ws, "A6", "수요기관", ", ".join(TARGET_ORG))
        write_meta(ws, "A7", "업무구분", service_text)
        write_meta(ws, "A8", "총 건수",  f"{len(df)}건")

        apply_excel_style(ws, start_row, {
            "A": 6,  "B": 12, "C": 10, "D": 52,
            "E": 32, "F": 12, "G": 20, "H": 12,
            "I": 10, "J": 12, "K": 18,
        }, number_cols=["K"])

    logger.info(f"  저장 완료: {file_path}")
    return file_path


# ══════════════════════════════════════════════════════════════════════
# 실행 단위
# ══════════════════════════════════════════════════════════════════════

def run() -> None:

    now = now_kst()

    # 월요일인 경우 이전 3일(금~일)
    if now.weekday() == 0:
        start_target = now - timedelta(days=3)   # 금요일
        end_target   = now - timedelta(days=1)   # 일요일

    # 화~금
    else:
        start_target = now - timedelta(days=1)
        end_target   = now - timedelta(days=1)

    # 수동 날짜가 설정된 경우 우선 적용
    start_date = globals().get("MANUAL_START") or start_target.strftime("%Y%m%d") + "0000"  # 전일 00:00
    end_date   = globals().get("MANUAL_END")   or end_target.strftime("%Y%m%d") + "2359"    # 전일 23:59
    
    mode = "수동" if globals().get("MANUAL_START") else "자동"
    logger.info(f"  날짜 모드  : [{mode}]")
    
    order_start_ym = shift_month(now, -2)                     # 당월 -2달
    order_end_ym   = shift_month(now, +10)                    # 당월 +10달

    logger.info(f"\n{'='*60}")
    logger.info(f"  실행 시각 : {now.strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info(
        f"  조회 기간 : "
        f"{start_date[:4]}-{start_date[4:6]}-{start_date[6:8]} "
        f"{start_date[8:10]}:{start_date[10:12]} "
        f"~ "
        f"{end_date[:4]}-{end_date[4:6]}-{end_date[6:8]} "
        f"{end_date[8:10]}:{end_date[10:12]}"
    )
    logger.info(f"  발주시기  : {order_start_ym} ~ {order_end_ym}")
    logger.info(f"  저장 폴더 : {OUTPUT_DIR}")
    logger.info(f"{'='*60}")

    run_time = now.strftime("%Y%m%d_%H%M")
    session = make_session()

    df_bid  = fetch_입찰공고(session, start_date, end_date)
    f_bid   = save_입찰공고(df_bid, start_date, end_date, run_time)

    df_spec = fetch_사전규격(session, start_date, end_date)
    f_spec  = save_사전규격(df_spec, start_date, end_date, run_time)

    df_plan = fetch_발주계획(session, start_date, end_date, order_start_ym, order_end_ym)
    f_plan  = save_발주계획(df_plan, start_date, end_date, run_time)

    logger.info(f"\n========== 완료 ==========")
    logger.info(f"  입찰공고 : {len(df_bid)}건  →  {f_bid}")
    logger.info(f"  사전규격 : {len(df_spec)}건  →  {f_spec}")
    logger.info(f"  발주계획 : {len(df_plan)}건  →  {f_plan}")


# ══════════════════════════════════════════════════════════════════════
# 스케줄러
# ══════════════════════════════════════════════════════════════════════

if __name__ == "__main__":

    now = now_kst()
    
    # 주말 체크
    if now.weekday() >= 5:
        logger.info("주말이므로 실행하지 않습니다.")
        sys.exit()
    
    current_hour = now.hour

    # 오전 8시 1회 실행
    run()
