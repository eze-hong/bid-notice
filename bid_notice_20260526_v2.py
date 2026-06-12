"""
나라장터 자동 조회 스크립트
  - 오전 08:00 실행 → 전일 0000 ~ 전일 2359
  - 발주계획 발주시기: 당월 -2달 ~ +10달
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

import math
import os
import time
import xml.etree.ElementTree as ET
import tempfile
import shutil
from contextlib import contextmanager
from pathlib import Path
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

import pandas as pd
import requests
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from dateutil.relativedelta import relativedelta
from dotenv import load_dotenv
from dataclasses import dataclass, field
import logging
logging.getLogger("urllib3").setLevel(logging.ERROR)


# ══════════════════════════════════════════════════════════════════════
# ★ 사용자 설정
# ══════════════════════════════════════════════════════════════════════
load_dotenv(r"C:\bid_notice\.env")

@dataclass(frozen=True)
class Config:
    service_key: str = field(
        default_factory=lambda: os.environ["SERVICE_KEY"]
    )
    # 엑셀 파일 저장 폴더
    output_dir: str = field(
        default_factory=lambda: os.environ["OUTPUT_DIR"]
    )
    log_root: str = r"C:\bid_notice\logs"
    rows: int = 50
    
    target_orgs: tuple = (
        "한국연구재단",
        "정보통신기획평가원",
        "한국과학기술정보연구원",
        "한국과학기술기획평가원",
        "중소기업기술정보진흥원",
        "한국보건산업진흥원",
        "정보통신산업진흥원",
        "한국과학기술연구원",
        "한국산업기술기획평가원",
        "한국지능정보사회진흥원",
        "교육부",
    )
    target_service_types: tuple = ("일반용역", "기술용역")
    
    manual_start: str | None = None
    manual_end:   str | None = None

CFG = Config();

# 수동 실행 시
#CFG = Config(manual_start="202605240000", manual_end="202605242359")


# ══════════════════════════════════════════════════════════════════════
# ★ 수동 날짜 설정
# ══════════════════════════════════════════════════════════════════════

#MANUAL_START = "202605240000"
#MANUAL_END   = "202605242359"


# ══════════════════════════════════════════════════════════════════════
# API 엔드포인트
# ══════════════════════════════════════════════════════════════════════

URL_BID  = "https://apis.data.go.kr/1230000/ad/BidPublicInfoService/getBidPblancListInfoServc"
URL_SPEC = "https://apis.data.go.kr/1230000/ao/HrcspSsstndrdInfoService/getPublicPrcureThngInfoServcPPSSrch"
URL_PLAN = "https://apis.data.go.kr/1230000/ao/OrderPlanSttusService/getOrderPlanSttusListServcPPSSrch"


# ══════════════════════════════════════════════════════════════════════
# 공통 유틸
# ══════════════════════════════════════════════════════════════════════
FONT_NAME    = "맑은 고딕"
HEADER_FILL  = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ROW_FILL_ODD = PatternFill(start_color="EEF2F9", end_color="EEF2F9", fill_type="solid")
THIN_SIDE    = Side(border_style="thin", color="C9C9C9")
CELL_BORDER  = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


def now_kst() -> datetime:
    return datetime.now(ZoneInfo("Asia/Seoul"))


def shift_month(dt: datetime, delta: int) -> str:
    """
    당월 기준으로 delta개월 이동한 YYYYMM 반환 (delta 음수=과거, 양수=미래)
    """
    return (dt + relativedelta(months=delta)).strftime("%Y%m")


def fmt_date(yyyymmddhhmi: str) -> str:
    d = yyyymmddhhmi
    return f"{d[:4]}/{d[4:6]}/{d[6:8]}" if len(d) >= 8 else d


def safe_int(value: str, default: int = 0) -> int:
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return default

@contextmanager
def atomic_write(target_path: Path):
    """
    임시 파일에 완전히 쓴 뒤 원자적으로 교체
    실패 시 임시 파일 자동 삭제
    """
    target_path = Path(target_path)
    target_path.parent.mkdir(parents=True, exist_ok=True)

    tmp_fd, tmp_path = tempfile.mkstemp(
        suffix=".xlsx",
        dir=target_path.parent
    )
    os.close(tmp_fd)

    try:
        yield tmp_path                       
        os.replace(tmp_path, target_path)
    except Exception:
        Path(tmp_path).unlink(missing_ok=True) 
        raise
        
os.system("chcp 65001")

#로그 설정
log_month = now_kst().strftime("%Y%m")
LOG_DIR = os.path.join(CFG.log_root, log_month)     #C:\bid_notice\logs\202605
os.makedirs(LOG_DIR, exist_ok=True)

log_time = now_kst().strftime("%Y%m%d_%H%M%S")  #20260512_164233.log
log_file = os.path.join(LOG_DIR, f"{log_time}.log")

logging.basicConfig(
level=logging.INFO,
format="%(asctime)s [%(levelname)s] %(message)s",
handlers=[
logging.FileHandler(log_file, encoding="utf-8"),
logging.StreamHandler()
]
)

logger = logging.getLogger(__name__)


def out_path(run_time: str, filename: str) -> str:
    month_dir = run_time[:6]
    
    save_dir = os.path.join(CFG.output_dir, month_dir, run_time)
    os.makedirs(save_dir, exist_ok=True)
    return os.path.join(save_dir, filename)

def make_session() -> requests.Session:
    session = requests.Session()
    retry = Retry(total=3, backoff_factor=1, status_forcelist=[500, 502, 503, 504])
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    return session


def fetch_all_pages(session: requests.Session, url: str, base_params: dict) -> list[ET.Element]:
    """페이지네이션을 순회하며 모든 <item> 요소 반환"""
    all_items: list[ET.Element] = []
    total_pages = 0
    try:
        params = {**base_params, "pageNo": 1}
        resp = session.get(url, params=params, timeout=(3, 15))
        resp.raise_for_status()

        root = ET.fromstring(resp.content)
        total_count = int(root.findtext(".//totalCount", default="0"))
        total_pages = math.ceil(total_count / base_params.get("numOfRows", 50))

        logger.info(f"  전체 건수: {total_count}  /  총 페이지: {total_pages}")\
        
        # 1페이지 바로 추가
        items = root.findall(".//item")
        if items:
            all_items.extend(items)
        
    except requests.exceptions.RequestException as e:
        logger.error(f"최초 API 연결 실패 (totalCount 조회 불가): {e}")
        raise

    # 2페이지부터 순회
    for page in range(2, total_pages + 1):
        logger.info(f"  [{page}/{total_pages}] 페이지 조회중...")
        params = {**base_params, "pageNo": page}

        try:
            resp = session.get(url, params=params, timeout=(3, 15))
            resp.raise_for_status()
            root  = ET.fromstring(resp.content)
            items = root.findall(".//item")

            if not items:
                print("  더 이상 데이터 없음, 순회 종료")
                break

            all_items.extend(items)

        except requests.exceptions.RequestException as e:
            logger.info(f"  [{page} 페이지] 수집 중 네트워크 에러 발생: {e}")
            raise RuntimeError(f"배치 수집 중 일부 테이더 유실 발생 (중단점: {page}page)") from e

        time.sleep(0.1)

    return all_items


def _filter_by_org_and_type(items, org_field, type_field):
    """
    입찰공고·사전규격용 org + service_type 필터
    """
    return [
        item for item in items
        if item.findtext(org_field, "") in CFG.target_orgs
        and item.findtext(type_field, "") in CFG.target_service_types
    ]

def _filter_by_org(items, org_field):
    """
    발주계획용 org 필터
    """
    return [
        item for item in items
        if item.findtext(org_field, "") in CFG.target_orgs
    ]

def write_meta(ws, cell: str, label: str, value) -> None:
    """
    메타 정보 작성 : 시작 셀에 label, 오른쪽 셀에 value 작성 & 스타일 적용
    """
    col_letter, row = coordinate_from_string(cell)
    col_idx = column_index_from_string(col_letter)
    next_col = get_column_letter(col_idx + 1)
    
    ws[cell].value = label
    ws[cell].font  = Font(name=FONT_NAME, bold=True, size=9, color="555555")
    
    value_cell = f"{next_col}{row}"
    ws[value_cell].value = value
    ws[value_cell].font  = Font(name=FONT_NAME, size=9)



def apply_excel_style(ws, start_row: int, column_widths: dict, number_cols: list = None) -> None:
    header_row = start_row + 1

    for cell in ws[header_row]:
        cell.font      = Font(name=FONT_NAME, bold=True, size=9, color="FFFFFF")
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = CELL_BORDER
    ws.row_dimensions[header_row].height = 30

    for i, row in enumerate(ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row)):
        fill = ROW_FILL_ODD if i % 2 == 0 else PatternFill()
        for cell in row:
            cell.font      = Font(name=FONT_NAME, size=9)
            cell.fill      = fill
            cell.border    = CELL_BORDER
            cell.alignment = Alignment(vertical="center")

    if number_cols:
        for col_letter in number_cols:
            col_idx = ord(col_letter.upper()) - 64
            for row in ws.iter_rows(
                min_row=header_row + 1, max_row=ws.max_row,
                min_col=col_idx, max_col=col_idx,
            ):
                for cell in row:
                    cell.number_format = "#,##0"
                    cell.alignment     = Alignment(horizontal="right", vertical="center")

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def write_no_data(ws, start_row: int, merge_cols: int = 5) -> None:
    """데이터 없을 때 안내 문구 삽입"""
    no_data_row = start_row + 2
    
    ws.merge_cells(start_row=no_data_row, start_column=1, end_row=no_data_row, end_column=5)
    cell = ws.cell(row=no_data_row, column=1)
    cell.value = "조회된 데이터가 없습니다."
    cell.font = Font(name=FONT_NAME, size=10, color="888888")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill      = PatternFill()
    cell.border    = Border()
    

# ══════════════════════════════════════════════════════════════════════
# 01. 입찰공고
# ══════════════════════════════════════════════════════════════════════

def fetch_입찰공고(session, start_date: str, end_date: str) -> pd.DataFrame:

    logger.info("\n===== 01. 입찰공고 조회 시작 =====")

    base_params = {
        "serviceKey": CFG.service_key,
        "numOfRows":  CFG.rows,
        "inqryDiv":   1,
        "inqryBgnDt": start_date,
        "inqryEndDt": end_date,
    }
    items = fetch_all_pages(session, URL_BID, base_params)
    
    latest_items = {}

    for item in items:
        bid_no = item.findtext("bidNtceNo", "")
        detail_bid_no = item.findtext("bidNtceOrd", "0")

        try:
            ord_no = int(detail_bid_no)
        except ValueError:
            ord_no = 0

        if bid_no not in latest_items:
            latest_items[bid_no] = item
        else:
            prev_ord = int(
                latest_items[bid_no].findtext("bidNtceOrd", "0")
            )

            if ord_no > prev_ord:
                latest_items[bid_no] = item

    
    filtered = _filter_by_org_and_type(
        latest_items.values(), "dminsttNm", "srvceDivNm"
    )
    
    results = []
    
    for item in filtered:
        
        bid_no = item.findtext("bidNtceNo", "")
        detail_bid_no = item.findtext("bidNtceOrd", "0")
        입찰공고번호 = f"{bid_no}-{detail_bid_no}"
        
        service_type = item.findtext("srvceDivNm", "")
        demand_org   = item.findtext("dminsttNm", "")

        results.append({
            "NO":           len(results) + 1,
            "업무구분":     service_type,
            "업무여부":     "",
            "구분":         item.findtext("ntceKindNm", ""),
            "입찰공고번호": 입찰공고번호,
            "공고명":       item.findtext("bidNtceNm", ""),
            "공고기관":     item.findtext("ntceInsttNm", ""),
            "수요기관":     demand_org,
            "게시일시":     item.findtext("bidNtceDt", ""),
            "입찰마감일시": item.findtext("bidClseDt", ""),
            #"단계":         "",
            #"세부절차":     "",
            #"세부절차상태": "",
            "계약체결방법명": item.findtext("cntrctCnclsMthdNm", ""),
            "낙찰방법명": item.findtext("sucsfbidMthdNm", ""),
            "배정예산금액": safe_int(item.findtext("asignBdgtAmt", "")),
            "추정가격": safe_int(item.findtext("presmptPrce", "")),
        })

    logger.info(f"  입찰공고 필터 결과: {len(results)}건")
    df = pd.DataFrame(results)

    if not df.empty:
        df = df.sort_values("게시일시", ascending=False).reset_index(drop=True)
        df["NO"] = df.index + 1
    return df


def save_입찰공고(df: pd.DataFrame, start_date: str, end_date: str, run_time: str,) -> str | None:
    ts        = now_kst().strftime("%Y%m%d_%H%M")
    file_path = out_path(run_time, f"01.입찰공고목록_{ts}.xlsx")
    start_row = 5

    with atomic_write(file_path) as tmp:
        with pd.ExcelWriter(tmp, engine="openpyxl") as writer:
            
            df.to_excel(writer, sheet_name="입찰공고", index=False, startrow=start_row)
            ws = writer.sheets["입찰공고"]
            
            ws["A1"] = "입찰공고목록"
            ws["A1"].font = Font(name=FONT_NAME, bold=True, size=15)

            write_meta(ws, "A2", "생성일",   now_kst().strftime("%Y/%m/%d %H:%M"))
            write_meta(ws, "A3", "조회기간", f"{fmt_date(start_date)} ~ {fmt_date(end_date)}")
            write_meta(ws, "A4", "총 건수",  f"{len(df)}건")
            
            if df.empty:
                write_no_data(ws, start_row, merge_cols=14)
            else:
                apply_excel_style(ws, start_row, {
                    "A": 6,  "B": 12, "C": 10, "D": 12, "E": 22,
                    "F": 46, "G": 28, "H": 26, "I": 20, "J": 20,
                    "K": 12, "L": 28, "M": 16, "N": 16,
                 }, number_cols=["M", "N"])
                

    logger.info(f"  저장 완료: {file_path}")
    return str(file_path)
   


# ══════════════════════════════════════════════════════════════════════
# 02. 사전규격공개
# ══════════════════════════════════════════════════════════════════════

def fetch_사전규격(session, start_date: str, end_date: str) -> pd.DataFrame:
    logger.info("\n===== 02. 사전규격공개 조회 시작 =====")

    base_params = {
        "serviceKey": CFG.service_key,
        "numOfRows":  CFG.rows,
        "inqryDiv":   1,
        "inqryBgnDt": start_date,
        "inqryEndDt": end_date,
    }
    items = fetch_all_pages(session, URL_SPEC, base_params)

    results, visited = [], set()
    
    unique_items = []
    for item in items:
        spec_no = item.findtext("bfSpecRgstNo", "")
        if spec_no in visited:
            continue
        visited.add(spec_no)
        unique_items.append(item)
    
    filtered = _filter_by_org_and_type(unique_items, "rlDminsttNm", "bsnsDivNm")
    
    for item in filtered:
        
        service_type = item.findtext("bsnsDivNm", "")
        demand_org   = item.findtext("rlDminsttNm", "")

        results.append({
            "NO":         len(results) + 1,
            "업무구분":   service_type,
            "업무여부":   "",
            "사업명":     item.findtext("prdctClsfcNoNm", ""),
            "수요기관":   demand_org,
            "공고기관":   item.findtext("orderInsttNm", ""),
            "담당자":     item.findtext("ofclNm", ""),
            "진행일자":   item.findtext("rcptDt", ""),
            "진행상태":   "",
            "참조여부":   "",
            "업체등록수": "",
            "예산금액":   safe_int(item.findtext("asignBdgtAmt", "0")),
        })

    logger.info(f"  사전규격 필터 결과: {len(results)}건")
    df = pd.DataFrame(results)
    if not df.empty:
        df = df.sort_values("진행일자", ascending=False).reset_index(drop=True)
        df["NO"] = df.index + 1
    return df


def save_사전규격(df: pd.DataFrame, start_date: str, end_date: str, run_time: str,) -> str | None:
    ts        = now_kst().strftime("%Y%m%d_%H%M")
    file_path = out_path(run_time, f"02.발주목록_사전규격공개_{ts}.xlsx")
    start_row = 8

    with atomic_write(file_path) as tmp:
        with pd.ExcelWriter(tmp, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="사전규격공개", index=False, startrow=start_row)
            ws = writer.sheets["사전규격공개"]

            ws["A1"] = "발주목록_사전규격공개"
            ws["A1"].font = Font(name=FONT_NAME, bold=True, size=15)

            write_meta(ws, "A2", "생성일",   now_kst().strftime("%Y-%m-%d %H:%M"))
            write_meta(ws, "A3", "검색유형", "사전규격공개")
            write_meta(ws, "A4", "수요기관", ", ".join(CFG.target_orgs))
            write_meta(ws, "A5", "업무구분", ", ".join(CFG.target_service_types))
            write_meta(ws, "A6", "진행일자", f"{fmt_date(start_date)} ~ {fmt_date(end_date)}")
            write_meta(ws, "A7", "총 건수",  f"{len(df)}건")
            
            if df.empty:
                write_no_data(ws, start_row, merge_cols=12)
            else:
                apply_excel_style(ws, start_row, {
                    "A": 6,  "B": 13, "C": 10, "D": 45,
                    "E": 24, "F": 24, "G": 11, "H": 20,
                    "I": 13, "J": 10, "K": 10, "L": 16,
                 }, number_cols=["L"])
                

    logger.info(f"  저장 완료: {file_path}")
    return str(file_path)



# ══════════════════════════════════════════════════════════════════════
# 03. 발주계획
# ══════════════════════════════════════════════════════════════════════

def fetch_발주계획(session, start_date: str, end_date: str,
                  order_start_ym: str, order_end_ym: str) -> pd.DataFrame:
    logger.info("\n===== 03. 발주계획 조회 시작 =====")

    base_params = {
        "serviceKey": CFG.service_key,
        "numOfRows":  CFG.rows,
        "inqryDiv":   1,
        "inqryBgnDt": start_date,
        "inqryEndDt": end_date,
        "orderBgnYm": order_start_ym,
        "orderEndYm": order_end_ym,
    }
    items = fetch_all_pages(session, URL_PLAN, base_params)

    results, visited = [], set()
    unique_items = []
    
    for item in items:
        plan_no = item.findtext("orderPlanUntyNo", "")
        if plan_no in visited:
            continue
        visited.add(plan_no)
        unique_items.append(item)
    
    filtered = _filter_by_org(unique_items, "orderInsttNm")
    
    for item in filtered:
        
        service_type = item.findtext("bsnsDivNm", "")
        demand_org   = item.findtext("orderInsttNm", "")
        발주년도      = item.findtext("orderYear", "")
        발주시기      = item.findtext("orderMnth", "")

        results.append({
            "No":       len(results) + 1,
            "업무구분": service_type,
            "업무여부": "",
            "사업명":   item.findtext("bizNm", ""),
            "수요기관": demand_org,
            "담당자":   item.findtext("ofclNm", ""),
            "진행일자": item.findtext("nticeDt", ""),
            "진행상태": "",
            "참조여부": "",
            "발주시기": f"{발주년도}/{발주시기.zfill(2)}",
            "예산금액": safe_int(item.findtext("sumOrderAmt", "0")),
        })

    logger.info(f"  발주계획 필터 결과: {len(results)}건")
    df = pd.DataFrame(results)
    if not df.empty:
        df = df.sort_values("진행일자", ascending=False).reset_index(drop=True)
        df["No"] = df.index + 1
    return df


def save_발주계획(df: pd.DataFrame, start_date: str, end_date: str, run_time: str,) -> str | None:
    ts        = now_kst().strftime("%Y%m%d_%H%M")
    file_path = out_path(run_time, f"03.발주목록_발주계획_{ts}.xlsx")
    start_row = 9

    month_list   = sorted(set(df["발주시기"].astype(str))) if not df.empty else []
    service_text = ", ".join(sorted(set(df["업무구분"].astype(str)))) if not df.empty else ""

    with atomic_write(file_path) as tmp:
       with pd.ExcelWriter(tmp, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="발주계획", index=False, startrow=start_row)
            ws = writer.sheets["발주계획"]

            ws["A1"] = "발주목록_발주계획"
            ws["A1"].font = Font(name=FONT_NAME, bold=True, size=15)

            write_meta(ws, "A2", "생성일",   now_kst().strftime("%Y-%m-%d %H:%M"))
            write_meta(ws, "A3", "검색유형", "발주계획")
            write_meta(ws, "A4", "진행일자", f"{fmt_date(start_date)} ~ {fmt_date(end_date)}")
            write_meta(ws, "A5", "발주시기", ", ".join(month_list))
            write_meta(ws, "A6", "수요기관", ", ".join(CFG.target_orgs))
            write_meta(ws, "A7", "업무구분", service_text)
            write_meta(ws, "A8", "총 건수",  f"{len(df)}건")


            if df.empty:
                write_no_data(ws, start_row, merge_cols=11)
            else:
                apply_excel_style(ws, start_row, {
                    "A": 6,  "B": 12, "C": 10, "D": 52,
                    "E": 32, "F": 12, "G": 20, "H": 12,
                    "I": 10, "J": 12, "K": 18,
                 }, number_cols=["K"])
                
                
    logger.info(f"  저장 완료: {file_path}")
    return str(file_path)
    


# ══════════════════════════════════════════════════════════════════════
# 실행 단위
# ══════════════════════════════════════════════════════════════════════

def calc_date_range(now: datetime) -> tuple[str, str]:
    """
    - 요일에 따라 조회 시작/종료 날짜 반환
    - 수동 설정 우선
    """
    if now.weekday() == 0:        # 월요일 -> 금~일
        start_target = now - timedelta(days=3)
        end_target   = now - timedelta(days=1)
    else:                       # 화~금 -> 전일
        start_target = now - timedelta(days=1)
        end_target   = now - timedelta(days=1)
        
    start = CFG.manual_start or start_target.strftime("%Y%m%d") + "0000"    # 전일 00:00
    end   = CFG.manual_end   or end_target.strftime("%Y%m%d")   + "2359"    # 전일 23:59
    
    return start, end
    
def run_all_fetches(
    session: requests.Session,
    start_date: str,
    end_date: str,
    order_start_ym: str,
    order_end_ym: str,
) -> dict[str, pd.DataFrame]:
    """
    세 API 순서대로 호출하고 DataFrame 딕셔너리로 반환
    """
    
    return {
        "입찰공고": fetch_입찰공고(session, start_date, end_date),
        "사전규격": fetch_사전규격(session, start_date, end_date),
        "발주계획": fetch_발주계획(session, start_date, end_date, order_start_ym, order_end_ym),
    }

def save_all(
    results: dict[str, pd.DataFrame],
    start_date: str,
    end_date: str,
    run_time: str,
) -> dict[str, str | None]:
    """
    DataFrame 딕셔너리를 받아 엑셀 저장 후 경로 딕셔너리 반환
    """
    return{
        "입찰공고": save_입찰공고(results["입찰공고"], start_date, end_date, run_time),
        "사전규격": save_사전규격(results["사전규격"], start_date, end_date, run_time),
        "발주계획": save_발주계획(results["발주계획"], start_date, end_date, run_time),
    }

def log_summary(results: dict[str, pd.DataFrame], paths: dict[str, str | None]) -> None:
    """
    수집, 저장 결과 로깅
    """
    logger.info(f"\n========== 완료 ==========")
    for key in results:
        count = len(results[key])
        path = paths.get(key) or "저장 없음"
        logger.info(f"  {key} : {count}건  →  {path}")

def run() -> None:
    try:
        now = now_kst()

        start_date, end_date = calc_date_range(now)
        order_start_ym       = shift_month(now, -2)                     # 당월 -2달
        order_end_ym         = shift_month(now, +10)                    # 당월 +10달
        run_time             = now.strftime("%Y%m%d_%H%M")
        
        
        mode = "수동" if CFG.manual_start else "자동"
        logger.info(f"  날짜 모드  : [{mode}]")
        logger.info(f"\n{'='*60}")
        logger.info(f"  실행 시각 : {now.strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info(
            f"  조회 기간 : "
            f"{start_date[:4]}-{start_date[4:6]}-{start_date[6:8]} "
            f"{start_date[8:10]}:{start_date[10:12]} "
            f"~ "
            f"{end_date[:4]}-{end_date[4:6]}-{end_date[6:8]} "
            f"{end_date[8:10]}:{end_date[10:12]}"
        )
        logger.info(f"  발주시기  : {order_start_ym} ~ {order_end_ym}")
        logger.info(f"  저장 폴더 : {CFG.output_dir}")
        logger.info(f"{'='*60}")

        session = make_session()
        results = run_all_fetches(session, start_date, end_date, order_start_ym, order_end_ym)
        paths   = save_all(results, start_date, end_date, run_time)
        log_summary(results, paths)
        
    except requests.RequestException as e:
        logger.error(f"네트워크 오류: {e}")
        raise
    except Exception as e:
        logger.critical(f"예기치 못한 오류: {e}", exc_info=True)
        raise


# ══════════════════════════════════════════════════════════════════════
# 스케줄러
# ══════════════════════════════════════════════════════════════════════

if __name__ == "__main__":

    now = now_kst()
    
    # 주말 체크
    if now.weekday() >= 5:
        logger.info("주말이므로 실행하지 않습니다.")
        sys.exit()
    
    current_hour = now.hour

    # 오전 8시 1회 실행
    run()
